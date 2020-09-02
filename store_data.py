from openpyxl import load_workbook
import constants as const

def update_excel(case_data, comment):
    workbook = load_workbook(const.file_path)
    worksheet = workbook.active
    max_row = worksheet.max_row
    excel_data = []
    excel_data.append(case_data['caseNumber'])
    excel_data.append(case_data['case_link'])
    excel_data.append(case_data['accountType'])
    excel_data.append(case_data['priority'])
    excel_data.append(case_data['subject'])
    excel_data.append(comment)
    worksheet.append(excel_data)
    workbook.save(const.file_path)







